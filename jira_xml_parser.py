# Copyright (c) 2020 Adrian S. Lemoine
#
#   Distributed under the Boost Software License, Version 1.0. 
#   (See accompanying file LICENSE_1_0.txt or copy at 
#   http://www.boost.org/LICENSE_1_0.txt)

import xml.etree.ElementTree as ET
from openpyxl import Workbook, load_workbook 
from openpyxl.utils import get_column_letter
import os
import sys
import argparse
from datetime import date, datetime

class Attributes:
    def __init__(self, lst, num):
        self.tag = lst[0]
        self.heading = lst[1]
        self.index = num
        self.column = num + 1

class JiraFields:
    def __init__(self, listoflist):
        self.field_list = []
        count = 0
        for el in listoflist:
            self.field_list.append(Attributes(el, count))
            count = count + 1
    def __getitem__(self, key):
        return self.field_list[key]
    def __len__(self):
        return len(self.field_list)
    def find_column(self, str):
        for el in self.field_list:
            if str == el.tag:
                return get_column_letter(el.column)
        raise NameError("No such tag found!")
    def find_column_num(self, str):
        for el in self.field_list:
            if str == el.tag:
                return el.column
        raise NameError("No such tag found!")
    def find_index(self, str):
        for el in self.field_list:
            if str == el.tag:
                return el.index
        raise NameError("No such tag found!")
    def headings(self):
        tmp = []
        for el in self.field_list:
            tmp.append(el.heading)
        return tmp

def read_xml(file):
    # Read in xml file
    if not os.path.exists(file):
        error_msg = "The file " + file + " could not be found!"
        sys.exit(error_msg)
    root = ET.parse(file).getroot()
    #print(root.tag)
    return root

def read_excel(file, jira_fields):
    # Set up Excel
    if os.path.exists(file):
        wb = load_workbook(filename = file)
    else:
       wb = Workbook()
       ws = wb.active
       # Add Headings
       ws.append(jira_fields.headings())
    return wb

def find_keys (root):
    # Finds all tickets in the XML file
    # The ticket's name is placed in the "key" XML tag
    key_ring = []
    for el in root.iter('key'):
        key_ring.append(el.text)
    return key_ring

def find_tag (key, tag):
    # Use the key to find a tags in the item that match the key
    # Returns a string
    tag_tablet = ''
    if tag is 'labels':
        tag_tablet = find_labels(key)
    elif tag is 'triage':
        tag_tablet = find_triage(key)
    elif tag is 'priority':
        tag_tablet = find_priority(key)
    elif tag is 'created' or tag is 'updated':
        tag_tablet = find_date(key, tag)
    elif tag is 'blocks':
        tag_tablet = find_blocks(key)
    elif tag is 'blocked_by':
        tag_tablet = find_blocked_by(key)
    else:
        text = "./channel/item/[key='"+key+"']"
        tag_tablet = xml_root.find(text).find(tag).text
    return tag_tablet

def find_labels(key):
    text = "./channel/item/[key='"+key+"']"
    labels = ""
    cntr = 0
    xelement = xml_root.find(text)
    for el in xelement.iter('label'):
       if cntr is 0:
           labels = el.text
       else:
           labels = labels + ', ' + el.text
       cntr = cntr + 1
    return labels

def find_triage(key):
    text = "./channel/item/customfields/..[key='"+key+"']/customfields/customfield[@id='customfield_14308']/customfieldvalues/"
    if ET.iselement(xml_root.find(text)):
        xelement = xml_root.find(text)
        return clean_string(xelement.text)
    else:
        return "None"

def find_priority(key):
    text = "./channel/item/[key='"+key+"']"
    prty = ""
    if ET.iselement(xml_root.find(text).find('priority')):
        prty = xml_root.find(text).find('priority').text
        return prty
    else:
        return "None"

def find_date(key, tag):
    text = "./channel/item/[key='"+key+"']"
    xelement = xml_root.find(text).find(tag).text 
    # Remove the " -0500" from the end of the string
    xelement = xelement[:-6]
    date_str = datetime.strptime(xelement,'%a, %d %b %Y %H:%M:%S').strftime("%m/%d/%Y %H:%M:%S")
    return date_str

def find_blocks(key):
    text = "./channel/item/issuelinks/..[key='"+key+"']/issuelinks/issuelinktype[@id='10000']/outwardlinks[@description='blocks']"
    blocks_issues = ""
    cntr = 0
    if ET.iselement(xml_root.find(text)):
        xelement = xml_root.find(text)
        for el in xelement.iter('issuekey'):
            if cntr is 0:
                blocks_issues = el.text
            else:
                blocks_issues = blocks_issues + ', ' + el.text
            cntr = cntr + 1
        return blocks_issues
    else:
        return ""

def find_blocked_by(key):
    text = "./channel/item/issuelinks/..[key='"+key+"']/issuelinks/issuelinktype[@id='10000']/inwardlinks[@description='is blocked by']"
    blocked_by = ""
    cntr = 0
    if ET.iselement(xml_root.find(text)):
        xelement = xml_root.find(text)
        for el in xelement.iter('issuekey'):
            if cntr is 0:
                blocked_by = el.text
            else:
                blocked_by = blocked_by + ', ' + el.text
            cntr = cntr + 1
        return blocked_by
    else:
        return ""

def clean_string(str):
    # Remove returns and excess whitespace.
    str = str.replace('\n',"").replace("  ", "")
    return str

def add_issue(issue):
    ws.append(issue)

# Program Start
## Set up arguments
parser = argparse.ArgumentParser(description = 
    "jira_xml_parser is a program which takes an XML file "
    "of Jira issues and inserts the new issues into a provided Excel "
    "file.")
parser.add_argument('xml_file', help='Path to Jira XML file.')
parser.add_argument('excel_file', nargs = '?', help = 'Path to Excel file.'
                    , default = 'jira_issues_'+date.today().strftime("%Y.%m.%d")+'.xlsx')
parser.add_argument('-f', '--force_update', action = 'store_true', 
                    help = 'Force all issues to be updated regardless of when it was last updated')

args = parser.parse_args()
xml_file = args.xml_file
excel_file = args.excel_file

print("Reading in ", xml_file)
print("Writing to ", excel_file)

# URL Root for Jira tickets
jira_url_root = 'http://ontrack-internal.amd.com/browse/'

# List the tags you are interested in collcecting and provide a heading
field_list = [["type", "Issue Type"]
             , ["key", "Key"]
             , ["summary", "Summary"]
             , ["assignee", "Assignee"]          
             , ["reporter", "Reporter"]
             , ["status", "Status"]
             , ["created", "Created"]
             , ["updated", "Updated"]
             , ["priority", "Priority"]
             , ["triage", "Triage Assignment"]
             , ["labels", "Lablels"]
             , ["blocks", "Blocks"]
             , ["blocked_by", "Blocked By"]]
            #, ["Target SW Release", "Target SW Release"]]

jira_fields = JiraFields(field_list)

xml_root = read_xml(xml_file)
wb = read_excel(excel_file, jira_fields)
ws = wb.active

key_column = ws[jira_fields.find_column("key")]
updated_column = ws[jira_fields.find_column("updated")]
excel_data = []
excel_keys = []
for el in range(1,len(key_column)):
    temp = []
    temp.append(el)
    temp.append(key_column[el].value)
    excel_keys.append(key_column[el].value)
    temp.append(updated_column[el].value)
    excel_data.append(temp)

# Find all unique tickets in XML
xml_keys = find_keys(xml_root)
# print(keys)

# Create a list of tickets and thier tags from the XML
xml_tickets = []
for el in xml_keys:
    temp = []
    for el2 in jira_fields:
        temp.append(find_tag(el, el2.tag))
    xml_tickets.append(temp)

# Write to Excel
## Search for copies
new_tickets = 0
updated_tickets = 0

# Index of the "updated" element
up_index = jira_fields.find_index("updated")

for el in range(len(xml_keys)):
    try:
        # Try to find each key in the excel keys.
        # Will throw an exception if it cannot find a match
       
        ## Return the index of the key
        result = excel_keys.index(xml_keys[el])
        ## Determine if all values should updated vased on "Updated" value
        ### Compare Updated value
        if xml_tickets[el][up_index] != excel_data[result][2].strftime("%m/%d/%Y %H:%M:%S") or args.force_update:
            updated_tickets += 1
            ## Update ticket with current information
            for el2 in range(len(jira_fields)):
                c = el2 + 1 # Columns start with 1
                r = result + 2 #Rows start with 1, add offset for heading
                ws.cell(row=r, column=c).value = xml_tickets[el][el2]
    except ValueError:
        ## The keys was not found append the new ticket
        new_tickets += 1
        add_issue(xml_tickets[el])

print("Updated Tickets: ", updated_tickets)
print("New Tickets", new_tickets)

# Add URLs to keys
#  For the "Key" Column
key_col = jira_fields.find_column_num("key")
for cell in ws.iter_rows(min_row=2, min_col=key_col, max_col=key_col):
    if cell[0].style is not "Hyperlink":
        iname = cell[0].value
        ws[cell[0].coordinate].hyperlink = jira_url_root + iname
        ws[cell[0].coordinate].style = "Hyperlink"

# Add Date format for created column
crt_col = jira_fields.find_column_num("created")
for cell in ws.iter_rows(min_row=2, min_col=crt_col, max_col=crt_col):
    if not cell[0].is_date:
        date_obj = datetime.strptime(cell[0].value, '%m/%d/%Y %H:%M:%S')
        ws[cell[0].coordinate].value = date_obj
        ws[cell[0].coordinate].number_format = 'MMM DD, YYYY'

# Add Date format for updated column
up_col = jira_fields.find_column_num("updated")
for cell in ws.iter_rows(min_row=2, min_col=up_col, max_col=up_col):
    if not cell[0].is_date:
        date_obj = datetime.strptime(cell[0].value, '%m/%d/%Y %H:%M:%S')
        ws[cell[0].coordinate].value = date_obj
        ws[cell[0].coordinate].number_format = 'MMM DD, YYYY'

wb.save(excel_file)